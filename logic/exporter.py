"""logic/exporter.py — Excel scorecard builder (T031).

MUST NOT import streamlit.
"""
import datetime
import io

import openpyxl

from logic.scoring import PASS_MARK


def build_scorecard(session_results: dict, historical_scorecard) -> bytes:
    """
    Build a 3-sheet Excel scorecard from session results.

    Sheets:
        1. Session Metadata — one data row with session summary.
        2. Question Results — one row per question (8 total).
        3. Category Summary — one row per TOGAF category in this session.

    The category_breakdown in session_results already contains cumulative
    values (merged with historical by logic.scoring.merge_historical before
    this function is called). The historical_scorecard parameter is accepted
    for API compatibility but is not used for computation here.

    Args:
        session_results:     The results dict from st.session_state.results.
        historical_scorecard: Accepted for signature compatibility; unused.

    Returns:
        Raw bytes of the .xlsx workbook (suitable for st.download_button).
    """
    wb = openpyxl.Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: Session Metadata
    # -----------------------------------------------------------------------
    ws_meta = wb.active
    ws_meta.title = "Session Metadata"
    ws_meta.append(
        [
            "user_name",
            "session_date",
            "total_score",
            "max_score",
            "pass_fail",
            "pass_mark",
            "time_limit_minutes",
        ]
    )
    ws_meta.append(
        [
            session_results["user_name"],
            datetime.date.today().isoformat(),
            session_results["total_score"],
            40,
            "PASS" if session_results["passed"] else "FAIL",
            PASS_MARK,
            90,
        ]
    )

    # -----------------------------------------------------------------------
    # Sheet 2: Question Results
    # -----------------------------------------------------------------------
    ws_qs = wb.create_sheet("Question Results")
    ws_qs.append(
        [
            "question_id",
            "scenario_snippet",
            "stem",
            "selected_option_id",
            "points_earned",
            "max_points",
            "category",
        ]
    )
    for pq in session_results["per_question"]:
        scenario_snippet = (pq["scenario"][:100] if pq["scenario"] else "")
        ws_qs.append(
            [
                pq["question_id"],
                scenario_snippet,
                pq["question"],
                pq["selected_option_id"] or "",
                pq["points_earned"],
                5,
                pq["primary_category"],
            ]
        )

    # -----------------------------------------------------------------------
    # Sheet 3: Category Summary
    # -----------------------------------------------------------------------
    ws_cat = wb.create_sheet("Category Summary")
    ws_cat.append(
        [
            "category",
            "session_points",
            "session_max",
            "session_pct",
            "cumulative_points",
            "cumulative_max",
            "cumulative_pct",
        ]
    )
    for cat, data in session_results["category_breakdown"].items():
        s_max = data["session_max"]
        s_pts = data["session_points"]
        c_max = data["cumulative_max"]
        c_pts = data["cumulative_points"]
        s_pct = 0.0 if s_max == 0 else round(s_pts / s_max * 100, 2)
        c_pct = 0.0 if c_max == 0 else round(c_pts / c_max * 100, 2)
        ws_cat.append(
            [cat, s_pts, s_max, s_pct, c_pts, c_max, c_pct]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

"""logic/importer.py — Excel scorecard parser (T032).

MUST NOT import streamlit.
"""
import openpyxl


def load_scorecard(file_obj) -> dict:
    """
    Parse a previously exported scorecard workbook.

    Validates that all three required sheets exist and that Category Summary
    contains the columns needed for cumulative tracking. Extra sheets and
    extra columns are silently ignored (forward-compatible).

    Args:
        file_obj: A file-like object (BytesIO or file handle) pointing to
                  a .xlsx workbook.

    Returns:
        {
            "category_summary": {
                "<category_name>": {
                    "cumulative_points": int,
                    "cumulative_max": int,
                }
            }
        }

    Raises:
        ValueError: If the file cannot be read or the schema is incompatible.
    """
    try:
        wb = openpyxl.load_workbook(file_obj)
    except Exception:
        raise ValueError("Incompatible scorecard schema: cannot read file")

    # Validate all required sheets are present
    required_sheets = ["Session Metadata", "Question Results", "Category Summary"]
    for name in required_sheets:
        if name not in wb.sheetnames:
            raise ValueError(
                f"Incompatible scorecard schema: missing sheet '{name}'"
            )

    # Validate required columns in Category Summary
    ws_cat = wb["Category Summary"]
    header_row = next(ws_cat.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h) if h is not None else "" for h in header_row]
    required_cols = ["category", "cumulative_points", "cumulative_max"]
    for col in required_cols:
        if col not in headers:
            raise ValueError(
                f"Incompatible scorecard schema: missing column '{col}'"
            )

    col_idx = {h: i for i, h in enumerate(headers)}

    # Parse category data rows
    cat_dict = {}
    for row in ws_cat.iter_rows(min_row=2, values_only=True):
        category = row[col_idx["category"]]
        if category is None:
            continue
        cat_dict[str(category)] = {
            "cumulative_points": int(row[col_idx["cumulative_points"]]),
            "cumulative_max": int(row[col_idx["cumulative_max"]]),
        }

    return {"category_summary": cat_dict}
